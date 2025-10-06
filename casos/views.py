from . import models
from . import forms

# Django Core Imports
from datetime import timedelta, date
import json
import os
from io import BytesIO
from collections import defaultdict
import calendar
from .microsoft_graph_service import enviar_email_graph, criar_pasta_caso, criar_subpastas
from notificacoes.servicos import preparar_notificacao # Vamos criar/renomear esta função
from .models import FluxoTrabalho, FluxoInterno as FluxoInternoModel
from django.contrib import messages
from django.shortcuts import redirect
from django.urls import reverse_lazy
# Django Imports
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q, Count, Sum # Adicionei o Sum aqui
from django.db.models.functions import TruncMonth
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.template.loader import render_to_string
from django.urls import reverse_lazy, reverse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from . import microsoft_graph_service
from django.views.generic import ListView, CreateView, DetailView, UpdateView, TemplateView, DeleteView
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.template.loader import render_to_string
from django.db.models import Sum
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill 
from weasyprint import HTML
from .models import Caso, DespesaCaso
from configuracoes.models import LogoConfig
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from .models import Caso
from .models import AcordoCaso, ParcelaAcordo
from .forms import AcordoCasoForm

# Third-Party Imports
import openpyxl
from openpyxl.styles import Font, Alignment
from weasyprint import HTML

# Local App Imports
from configuracoes.models import LogoConfig
from notificacoes.servicos import enviar_notificacao

# --- Imports do App 'casos' ---
from .microsoft_graph_service import (
    criar_pasta_caso, criar_subpastas, listar_arquivos_e_pastas, 
    upload_arquivo, deletar_item, criar_nova_pasta, obter_url_preview,
    enviar_email as enviar_email_service
)

from .models import (
    Advogado, AndamentoCaso, AcaoEtapa, Caso, Cliente, Campo, EmailCaso, 
    EmailTemplate, EstruturaPasta, EtapaFluxo, FluxoInterno as FluxoInternoModel, 
    FluxoTrabalho, HistoricoEtapa, InstanciaAcao, OpcaoDecisao, Produto, Status, 
    Timesheet, UserSignature, ValorCampoCaso,
    # Importando os novos modelos
    DespesaCaso, AcordoCaso, ParcelaAcordo
)

from .forms import (
    AndamentoCasoForm, CasoCreateForm, CasoUpdateForm, EnviarEmailForm, 
    FluxoInternoForm, LancamentoHorasForm, DespesaCasoForm
)

from .tasks import buscar_detalhes_email_enviado, processar_email_webhook

Usuario = get_user_model()


def _mudar_etapa_fluxo(request, caso, nova_etapa):
    from django.contrib import messages
    etapa_antiga = caso.etapa_atual
    
    if etapa_antiga:
        historico = models.HistoricoEtapa.objects.filter(caso=caso, etapa=etapa_antiga, data_saida__isnull=True).first()
        if historico:
            historico.data_saida = timezone.now()
            historico.save()
            tempo_gasto = (historico.data_saida - historico.data_entrada).days
            models.FluxoInterno.objects.create(
                caso=caso, 
                data_fluxo=timezone.now().date(), 
                descricao=f"[WORKFLOW] Etapa '{etapa_antiga.nome}' finalizada. Tempo: {tempo_gasto} dias.", 
                usuario_criacao=request.user
            )
    
    caso.etapa_atual = nova_etapa
    caso.save(update_fields=['etapa_atual'])

    if nova_etapa:
        models.HistoricoEtapa.objects.create(caso=caso, etapa=nova_etapa)
        models.FluxoInterno.objects.create(
            caso=caso, 
            data_fluxo=timezone.now().date(), 
            descricao=f"[WORKFLOW] Caso entrou na etapa: '{nova_etapa.nome}'.", 
            usuario_criacao=request.user
        )
        
        count_acoes = 0
        for acao_modelo in nova_etapa.acoes.all():
            responsavel_final = None
            if acao_modelo.tipo_responsavel == 'CRIADOR_ACAO':
                responsavel_final = request.user
            elif acao_modelo.tipo_responsavel == 'RESPONSAVEL_CASO':
                if caso.advogado_responsavel:
                    responsavel_final = caso.advogado_responsavel.user
            elif acao_modelo.tipo_responsavel == 'USUARIO_FIXO':
                responsavel_final = acao_modelo.responsavel_fixo

            if not responsavel_final:
                responsavel_final = request.user
            
            prazo_final = None
            if acao_modelo.prazo_dias > 0:
                start_date = timezone.now()
                dias_a_adicionar = acao_modelo.prazo_dias
                if acao_modelo.tipo_prazo == 'uteis':
                    dias_extras = (dias_a_adicionar // 5) * 2
                    prazo_final = start_date + timedelta(days=dias_a_adicionar + dias_extras)
                else: 
                    prazo_final = start_date + timedelta(days=dias_a_adicionar)
            
            models.InstanciaAcao.objects.create(
                caso=caso, 
                acao_modelo=acao_modelo, 
                responsavel=responsavel_final, 
                prazo_final=prazo_final
            )
            count_acoes += 1
        
        if count_acoes > 0:
            messages.info(request, f"{count_acoes} nova(s) ação(ões) criada(s) para a etapa '{nova_etapa.nome}'.")



def get_campos_for_produto_ajax(request):
    produto_id, caso_id = request.GET.get('produto_id'), request.GET.get('caso_id')
    if not produto_id: return JsonResponse([], safe=False)
    campos_data = []
    campos = Campo.objects.filter(produtos__id=produto_id).order_by('nome_label')
    for campo in campos:
        valor_existente = ''
        if caso_id:
            try:
                valor_obj = ValorCampoCaso.objects.get(caso_id=caso_id, campo=campo)
                valor_existente = valor_obj.valor
            except ValorCampoCaso.DoesNotExist: pass
        campos_data.append({'nome_label': campo.nome_label, 'nome_tecnico': campo.nome_tecnico, 'tipo_campo': campo.tipo_campo, 'valor': valor_existente})
    return JsonResponse(campos_data, safe=False)

class CasoListView(LoginRequiredMixin, ListView):
    model = Caso
    template_name = 'casos/caso_list.html'
    context_object_name = 'casos'
    paginate_by = 15
    def get_queryset(self): return get_casos_filtrados(self.request)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        query_params = self.request.GET.copy()
        if 'page' in query_params: del query_params['page']
        context['querystring'] = query_params.urlencode()
        context['clientes_list'], context['produtos_list'], context['status_list'] = Cliente.objects.all().order_by('nome_razao_social'), Produto.objects.all().order_by('nome'), Status.objects.all().order_by('nome')
        return context

class CasoCreateView(LoginRequiredMixin, CreateView):
    model = Caso
    form_class = CasoCreateForm
    template_name = 'casos/caso_form_create.html'

    def form_valid(self, form):
        # Salva o objeto do formulário, mas não o envia para o banco ainda
        self.object = form.save(commit=False)
        
        # Lógica para iniciar o workflow
        try:
            fluxo = FluxoTrabalho.objects.get(cliente=self.object.cliente, produto=self.object.produto)
            primeira_etapa = fluxo.etapas.order_by('ordem').first()
            if primeira_etapa:
                self.object.etapa_atual = primeira_etapa
            else:
                messages.warning(self.request, "Fluxo de trabalho aplicável não possui etapas.")
        except FluxoTrabalho.DoesNotExist:
            messages.warning(self.request, "Nenhuma regra de fluxo de trabalho encontrada para este cliente e produto.")
        
        # Agora salva o objeto no banco para obter um ID
        self.object.save()
        
        # Inicia o fluxo se uma etapa foi definida
        if self.object.etapa_atual:
            _mudar_etapa_fluxo(self.request, self.object, self.object.etapa_atual)
            messages.info(self.request, f"Caso iniciado no fluxo '{self.object.etapa_atual.fluxo_trabalho.nome}'.")

        # Lógica de integração com o SharePoint
        try:
            nome_pasta_sanitizado = str(self.object.id)
            pasta_criada_json = criar_pasta_caso(nome_pasta_sanitizado)
            if pasta_criada_json:
                self.object.sharepoint_folder_id = pasta_criada_json['id']
                self.object.sharepoint_folder_url = pasta_criada_json['webUrl']
                self.object.save(update_fields=['sharepoint_folder_id', 'sharepoint_folder_url'])
                
                id_pasta_pai = pasta_criada_json['id']
                subpastas = [p.nome_pasta for p in self.object.produto.estrutura_pastas.all()]
                if subpastas:
                    subpastas_sanitizadas = ["".join(c if c not in r'<>:"/\|?*' else '-' for c in nome_sub) for nome_sub in subpastas]
                    criar_subpastas(id_pasta_pai, subpastas_sanitizadas)
                messages.success(self.request, "Pasta do caso criada no SharePoint.")
            else:
                messages.warning(self.request, "O caso foi criado, mas não foi possível criar a pasta no SharePoint.")
        except Exception as e:
            messages.error(self.request, f"Falha na integração com SharePoint: {e}")

        # --- NOVA LÓGICA DE ENVIO DE E-MAIL VIA MICROSOFT GRAPH ---
        try:
            # 1. Prepara o conteúdo do e-mail (assunto, corpo, destinatários)
            contexto_notificacao = {'caso': self.object, 'usuario_acao': self.request.user}
            sucesso_preparacao, dados_email = preparar_notificacao(slug_evento='novo-caso-criado', contexto=contexto_notificacao)

            if sucesso_preparacao:
                # 2. Envia o e-mail usando a conta do usuário logado
                sucesso_envio, mensagem_envio = enviar_email_graph(
                    usuario_remetente=self.request.user,
                    destinatarios=dados_email['destinatarios'],
                    assunto=dados_email['assunto'],
                    corpo_html=dados_email['corpo']
                )

                if sucesso_envio:
                    messages.info(self.request, "Notificação de novo caso enviada com sucesso pela sua conta Microsoft.")
                    # 3. Registra a ação no Fluxo Interno
                    FluxoInternoModel.objects.create(
                        caso=self.object,
                        data_fluxo=timezone.now().date(),
                        descricao=f"[SISTEMA] E-mail de 'Novo Caso Criado' enviado para: {', '.join(dados_email['destinatarios'])}.",
                        usuario_criacao=self.request.user
                    )
                else:
                    messages.error(self.request, f"Falha ao enviar notificação via Microsoft: {mensagem_envio}")
            else:
                # 'dados_email' aqui contém a mensagem de erro da preparação
                messages.warning(self.request, f"Não foi possível preparar as notificações: {dados_email}")

        except Exception as e:
            messages.error(self.request, f"Ocorreu um erro inesperado ao tentar enviar as notificações: {e}")
        # --- FIM DA NOVA LÓGICA ---
            
        return redirect(self.get_success_url())

    def get_success_url(self):
        # Redireciona para a página de UPDATE, que é mais útil após a criação
        return reverse_lazy('casos:caso_update', kwargs={'pk': self.object.pk})

class CasoUpdateView(LoginRequiredMixin, UpdateView):
    model = Caso
    form_class = CasoUpdateForm
    template_name = 'casos/caso_form_update.html'
    def get_success_url(self): return reverse_lazy('casos:caso_detail', kwargs={'pk': self.object.pk})

class CasoDetailView(LoginRequiredMixin, DetailView):
    model = models.Caso
    template_name = 'casos/caso_detail.html'
    context_object_name = 'caso'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        caso = self.get_object()
        
        # --- BUSCA OS DADOS PARA AS NOVAS ABAS ---
        # Busca todas as despesas relacionadas a este caso, ordenadas pela data mais recente
        context['despesas_do_caso'] = caso.despesas.all().order_by('-data_despesa')
        
        # Busca todos os acordos e, para cada um, já busca as parcelas relacionadas
        # O prefetch_related é uma otimização para evitar múltiplas consultas ao banco
        context['acordos_do_caso'] = caso.acordos.prefetch_related('parcelas').order_by('-data_acordo')

        # Calcula o total de despesas (exemplo de como fazer um cálculo)
        context['total_despesas'] = caso.despesas.aggregate(total=Sum('valor'))['total'] or 0.00
        # -------------------------------------------

        # --- Mantém toda a sua lógica existente ---
        context['arquivos_sharepoint'] = []
        if caso.sharepoint_folder_id:
            context['arquivos_sharepoint'] = microsoft_graph_service.listar_arquivos_e_pastas(caso.sharepoint_folder_id)
        
        email_form = forms.EnviarEmailForm()
        templates = models.EmailTemplate.objects.all()
        assinaturas = models.UserSignature.objects.filter(usuario=self.request.user)
        
        email_form.fields['modelo_id'].choices = [('', '---------')] + [(t.id, t.nome) for t in templates]
        email_form.fields['assinatura_id'].choices = [('', 'Nenhuma')] + [(a.id, a.nome) for a in assinaturas]
        
        context['email_form'] = email_form
        context['email_templates_json'] = json.dumps({t.id: {'assunto': t.assunto, 'corpo': t.corpo} for t in templates})
        context['user_signatures_json'] = json.dumps({a.id: a.corpo_html for a in assinaturas})
        context['emails_do_caso'] = models.EmailCaso.objects.filter(caso=caso)
        
        context['fluxo_interno_form'] = forms.FluxoInternoForm()
        context['andamento_caso_form'] = forms.AndamentoCasoForm()
        context['lancamento_horas_form'] = LancamentoHorasForm()
        
        context['acoes_pendentes'] = models.InstanciaAcao.objects.filter(caso=caso, status='P').order_by('data_criacao')
        context['acoes_concluidas'] = models.InstanciaAcao.objects.filter(caso=caso, status='C').order_by('-data_conclusao')
        context['historico_etapas'] = models.HistoricoEtapa.objects.filter(caso=caso).order_by('data_entrada')
        
        return context

def get_acoes_filtradas(request):
    """Função auxiliar que busca e filtra as InstanciaAcao."""
    queryset = models.InstanciaAcao.objects.select_related(
        'caso__cliente', 'acao_modelo', 'responsavel'
    ).order_by('prazo_final')

    # Filtro por Texto
    texto = request.GET.get('texto', '')
    if texto:
        query_texto = (
            Q(acao_modelo__titulo__icontains=texto) |
            Q(caso__titulo_caso__icontains=texto) |
            Q(caso__valores_dinamicos__valor__icontains=texto)
        )
        if texto.isdigit():
            query_texto |= Q(caso__id=texto)
        queryset = queryset.filter(query_texto).distinct()

    # Outros Filtros
    if responsavel_id := request.GET.get('responsavel', ''):
        queryset = queryset.filter(responsavel_id=responsavel_id)
    
    if status := request.GET.get('status', ''):
        queryset = queryset.filter(status=status)

    if cliente_id := request.GET.get('cliente', ''):
        queryset = queryset.filter(caso__cliente_id=cliente_id)

    if prazo_de := request.GET.get('prazo_de', ''):
        queryset = queryset.filter(prazo_final__date__gte=prazo_de)

    if prazo_ate := request.GET.get('prazo_ate', ''):
        queryset = queryset.filter(prazo_final__date__lte=prazo_ate)
        
    return queryset

class AcaoListView(LoginRequiredMixin, ListView):
    model = models.InstanciaAcao
    template_name = 'casos/acao_list.html'
    context_object_name = 'acoes'
    paginate_by = 20

    def get_queryset(self):
        return get_acoes_filtradas(self.request)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = "Gerenciador de Ações"
        
        # Para os filtros do formulário
        context['responsaveis'] = Usuario.objects.filter(is_active=True).order_by('first_name', 'last_name')
        context['clientes'] = models.Cliente.objects.all().order_by('nome_razao_social')
        context['status_choices'] = models.InstanciaAcao.STATUS_CHOICES
        
        # Para manter os filtros na paginação
        query_params = self.request.GET.copy()
        if 'page' in query_params:
            del query_params['page']
        context['querystring'] = query_params.urlencode()
        
        return context

@login_required
def exportar_acoes_excel(request):
    acoes_filtradas = get_acoes_filtradas(request)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="relatorio_acoes.xlsx"'
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Relatório de Ações'

    headers = ['ID Caso', 'Título do Caso', 'Ação', 'Responsável', 'Prazo Final', 'Status']
    sheet.append(headers)
    for cell in sheet[1]: cell.font = Font(bold=True)

    for acao in acoes_filtradas:
        sheet.append([
            acao.caso.id,
            acao.caso.titulo_caso,
            acao.acao_modelo.titulo,
            acao.responsavel.get_full_name() if acao.responsavel else 'N/A',
            acao.prazo_final.strftime('%d/%m/%Y') if acao.prazo_final else 'Sem prazo',
            acao.get_status_display(),
        ])
    workbook.save(response)
    return response


@login_required
def add_andamento_caso(request, caso_pk):
    caso = get_object_or_404(Caso, pk=caso_pk)
    if request.method == 'POST':
        form = AndamentoCasoForm(request.POST)
        if form.is_valid():
            andamento = form.save(commit=False)
            andamento.caso, andamento.usuario_criacao = caso, request.user
            andamento.save()
            messages.success(request, "Andamento adicionado.")
    return redirect(reverse('casos:caso_detail', kwargs={'pk': caso_pk}) + '#andamento-caso-tab-pane')

def get_casos_filtrados(request):
    queryset = Caso.objects.select_related('cliente', 'produto', 'status', 'etapa_atual').order_by('-id')
    titulo = request.GET.get('titulo', '')
    if titulo:
        query_texto = (Q(titulo_caso__icontains=titulo) | Q(valores_dinamicos__valor__icontains=titulo, valores_dinamicos__campo__nome_tecnico='aviso') | Q(valores_dinamicos__valor__icontains=titulo, valores_dinamicos__campo__nome_tecnico='segurado'))
        query_final = query_texto | Q(id=titulo) if titulo.isdigit() else query_texto
        queryset = queryset.filter(query_final).distinct()
    if cliente_id := request.GET.get('cliente', ''): queryset = queryset.filter(cliente_id=cliente_id)
    if produto_id := request.GET.get('produto', ''): queryset = queryset.filter(produto_id=produto_id)
    if status_id := request.GET.get('status', ''): queryset = queryset.filter(status_id=status_id)
    return queryset

class CasoPesquisaView(LoginRequiredMixin, ListView):
    model = Caso
    template_name = 'casos/caso_pesquisa.html'
    context_object_name = 'casos'
    paginate_by = 25
    def get_queryset(self): return get_casos_filtrados(self.request)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'], context['clientes'], context['advogados'], context['status_list'] = "Pesquisa Avançada", Cliente.objects.all(), Advogado.objects.all(), Status.objects.all()
        return context

@login_required
def add_fluxo_interno(request, caso_pk):
    caso = get_object_or_404(Caso, pk=caso_pk)
    if request.method == 'POST':
        form = FluxoInternoForm(request.POST)
        if form.is_valid():
            fluxo = form.save(commit=False)
            fluxo.caso, fluxo.usuario_criacao = caso, request.user
            fluxo.save()
            messages.success(request, "Registro adicionado ao Fluxo Interno.")
    return redirect(reverse('casos:caso_detail', kwargs={'pk': caso_pk}) + '#fluxo-interno-tab-pane')

@login_required
def add_timesheet(request, caso_pk):
    caso = get_object_or_404(Caso, pk=caso_pk)
    if request.method == 'POST':
        form = TimesheetForm(request.POST)
        if form.is_valid():
            timesheet = form.save(commit=False)
            timesheet.caso = caso
            tempo_str = form.cleaned_data['tempo_str']
            horas, minutos = map(int, tempo_str.split(':'))
            timesheet.tempo = timedelta(hours=horas, minutes=minutos)
            timesheet.save()
    return redirect(reverse('casos:caso_detail', kwargs={'pk': caso_pk}) + '#timesheet-tab-pane')

@login_required
def exportar_casos_excel(request):
    casos_filtrados = get_casos_filtrados(request)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="relatorio_casos.xlsx"'
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Relatório de Casos'
    headers = ['ID', 'Título', 'Cliente', 'Produto', 'Status', 'Etapa Atual']
    sheet.append(headers)
    for cell in sheet[1]: cell.font = Font(bold=True)
    for caso in casos_filtrados:
        sheet.append([caso.id, caso.titulo_caso, str(caso.cliente), str(caso.produto), str(caso.status), caso.etapa_atual.nome if caso.etapa_atual else "-"])
    workbook.save(response)
    return response

@login_required
def exportar_andamentos_excel(request, caso_pk):
    caso = get_object_or_404(models.Caso, pk=caso_pk)
    andamentos = caso.andamentos_caso.all()
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="andamentos_caso_{caso.id}.xlsx"'
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Andamentos'
    sheet.merge_cells('A1:C1')
    titulo_cell = sheet['A1']
    titulo_cell.value = f"Relatório de Andamentos - Caso: {caso.titulo_caso}"
    titulo_cell.font = Font(bold=True, size=14)
    titulo_cell.alignment = Alignment(horizontal='center')
    sheet.append([])
    headers = ['Data', 'Usuário', 'Descrição']
    sheet.append(headers)
    for cell in sheet[3]:
        cell.font = Font(bold=True)
    for andamento in andamentos:
        sheet.append([
            andamento.data_andamento.strftime("%d/%m/%Y %H:%M"),
            andamento.usuario_criacao.get_full_name() if andamento.usuario_criacao else "Sistema",
            andamento.descricao
        ])
    workbook.save(response)
    return response

@login_required
def exportar_andamentos_pdf(request, caso_pk):
    caso = get_object_or_404(models.Caso, pk=caso_pk)
    andamentos = caso.andamentos_caso.all()
    logo_url_completa = None
    try:
        logo_config = LogoConfig.objects.get(ativo=True)
        if logo_config.logo:
            logo_url_completa = request.build_absolute_uri(logo_config.logo.url)
    except LogoConfig.DoesNotExist:
        pass
    context = {'caso': caso, 'andamentos': andamentos, 'logo_url': logo_url_completa}
    html_string = render_to_string('casos/pdf/andamento_pdf_template.html', context)
    base_url = request.build_absolute_uri('/')
    html = HTML(string=html_string, base_url=base_url)
    pdf = html.write_pdf()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="andamentos_caso_{caso.id}.pdf"'
    return response

@login_required
def exportar_timesheet_excel(request, caso_pk):
    caso = get_object_or_404(models.Caso, pk=caso_pk)
    timesheets = caso.timesheets.all().order_by('data_execucao')
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="timesheet_caso_{caso.id}.xlsx"'
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Lançamentos de Horas'
    sheet.merge_cells('A1:D1')
    titulo_cell = sheet['A1']
    titulo_cell.value = f"Relatório de Horas - Caso: {caso.titulo_caso}"
    titulo_cell.font = Font(bold=True, size=14)
    titulo_cell.alignment = Alignment(horizontal='center')
    sheet.row_dimensions[1].height = 20
    headers = ['Data de Execução', 'Profissional', 'Tempo Gasto (HH:MM)', 'Descrição da Atividade']
    sheet.append([])
    sheet.append(headers)
    for cell in sheet[3]:
        cell.font = Font(bold=True)
    sheet.column_dimensions['A'].width, sheet.column_dimensions['B'].width, sheet.column_dimensions['C'].width, sheet.column_dimensions['D'].width = 18, 30, 20, 60
    for ts in timesheets:
        data_formatada = ts.data_execucao.strftime("%d/%m/%Y")
        horas, minutos = ts.tempo.seconds // 3600, (ts.tempo.seconds % 3600) // 60
        tempo_formatado = f"{horas:02d}:{minutos:02d}"
        sheet.append([
            data_formatada,
            ts.profissional.get_full_name() or ts.profissional.username,
            tempo_formatado,
            ts.descricao
        ])
    if timesheets:
        proxima_linha = sheet.max_row + 1
        total_label_cell = sheet.cell(row=proxima_linha, column=2)
        total_label_cell.value, total_label_cell.font, total_label_cell.alignment = "Total de Horas:", Font(bold=True), Alignment(horizontal='right')
        total_value_cell = sheet.cell(row=proxima_linha, column=3)
        total_value_cell.value, total_value_cell.font = caso.total_horas_trabalhadas, Font(bold=True)
    workbook.save(response)
    return response

@login_required
def exportar_timesheet_pdf(request, caso_pk):
    caso = get_object_or_404(models.Caso, pk=caso_pk)
    timesheets = caso.timesheets.all().order_by('data_execucao')
    logo_url_completa = None
    try:
        logo_config = LogoConfig.objects.get(ativo=True)
        if logo_config.logo:
            logo_url_completa = request.build_absolute_uri(logo_config.logo.url)
    except LogoConfig.DoesNotExist:
        pass

    context = {
        'caso': caso, 
        'timesheets': timesheets, 
        'total_horas': caso.total_horas_trabalhadas, 
        'logo_url': logo_url_completa
    }

    html_string = render_to_string('casos/pdf/timesheet_pdf_template.html', context)
    base_url = request.build_absolute_uri('/')
    html = HTML(string=html_string, base_url=base_url)
    pdf = html.write_pdf()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="timesheet_caso_{caso.id}.pdf"'
    return response
@login_required
def criar_pasta_anexo_view(request, caso_pk):
    if request.method == 'POST':
        caso, nome_nova_pasta = get_object_or_404(Caso, pk=caso_pk), request.POST.get('nome_pasta')
        if nome_nova_pasta and caso.sharepoint_folder_id:
            nome_pasta_sanitizado = "".join(c if c not in r'<>:"/\|?*' else '-' for c in nome_nova_pasta)
            if criar_nova_pasta(caso.sharepoint_folder_id, nome_pasta_sanitizado): messages.success(request, f"Pasta '{nome_pasta_sanitizado}' criada.")
            else: messages.error(request, "Falha ao criar a pasta.")
    return redirect(reverse('casos:caso_detail', kwargs={'pk': caso_pk}) + '#anexos-sharepoint-tab-pane')

@login_required
def upload_arquivo_anexo_view(request, caso_pk):
    if request.method == 'POST':
        caso, arquivos_upload, id_pasta_pai = get_object_or_404(Caso, pk=caso_pk), request.FILES.getlist('arquivo'), request.POST.get('parent_folder_id', caso.sharepoint_folder_id)
        if not arquivos_upload: messages.warning(request, "Nenhum arquivo selecionado.")
        if arquivos_upload and id_pasta_pai:
            sucessos, falhas = 0, 0
            for arquivo in arquivos_upload:
                if upload_arquivo(id_pasta_pai, arquivo.name, arquivo.read()): sucessos += 1
                else:
                    falhas += 1
                    messages.error(request, f"Falha ao enviar: {arquivo.name}")
            if sucessos > 0: messages.success(request, f"{sucessos} arquivo(s) enviado(s).")
            if falhas == 0 and sucessos == 0: messages.error(request, "Falha no envio dos arquivos.")
    return redirect(reverse('casos:caso_detail', kwargs={'pk': caso_pk}) + '#anexos-sharepoint-tab-pane')

@login_required
def listar_subpasta_ajax(request, folder_id):
    if request.method == 'GET': return JsonResponse(listar_arquivos_e_pastas(folder_id), safe=False)
    return JsonResponse([], safe=False)

@login_required
def preview_arquivo_view(request, item_id):
    preview_url = obter_url_preview(item_id)
    if preview_url: return redirect(preview_url)
    messages.error(request, "Não foi possível gerar o link de visualização.")
    return redirect(request.META.get('HTTP_REFERER', '/'))

@login_required
def deletar_item_view(request, caso_pk, item_id):
    if request.method == 'POST':
        if deletar_item(item_id): messages.success(request, "Item deletado.")
        else: messages.error(request, "Falha ao deletar o item.")
    return redirect(reverse('casos:caso_detail', kwargs={'pk': caso_pk}) + '#anexos-sharepoint-tab-pane')

@login_required
def enviar_email_view(request, caso_pk):
    caso = get_object_or_404(Caso, pk=caso_pk)
    if request.method == 'POST':
        form = EnviarEmailForm(request.POST)
        templates, assinaturas = EmailTemplate.objects.all(), UserSignature.objects.filter(usuario=request.user)
        form.fields['modelo_id'].choices, form.fields['assinatura_id'].choices = [('', '---------')] + [(t.id, t.nome) for t in templates], [('', 'Nenhuma')] + [(a.id, a.nome) for a in assinaturas]
        if form.is_valid():
            para, assunto, corpo, remetente_email = form.cleaned_data['para'], form.cleaned_data['assunto'], form.cleaned_data['corpo'], request.user.email
            if enviar_email_service(remetente_email, para, assunto, corpo):
                email_caso_obj = EmailCaso.objects.create(caso=caso, microsoft_message_id=f"enviado_{timezone.now().timestamp()}", de=remetente_email, para=para, assunto=assunto, preview=corpo[:255], corpo_html=corpo, data_envio=timezone.now(), is_sent=True)
                messages.success(request, "E-mail enviado e registrado.")
                buscar_detalhes_email_enviado.delay(remetente_email, email_caso_obj.id, para, assunto)
            else: messages.error(request, "Falha ao enviar o e-mail.")
        else: messages.error(request, f"Formulário inválido: {form.errors.as_text()}")
    return redirect(reverse('casos:caso_detail', kwargs={'pk': caso_pk}) + '#email-tab-pane')

@csrf_exempt
def microsoft_graph_webhook(request):
    if 'validationToken' in request.GET: return HttpResponse(request.GET['validationToken'], content_type='text/plain', status=200)
    if request.method == 'POST':
        try:
            notification_data = json.loads(request.body)
            for notification in notification_data.get('value', []):
                processar_email_webhook.delay(notification['subscriptionId'], notification['resourceData']['id'])
            return HttpResponse(status=202)
        except Exception as e:
            print(f"ERRO ao processar webhook: {e}")
            return HttpResponse(status=400)
    return HttpResponse("Endpoint de Webhook. Acesso inválido.", status=405)

def add_generic_ajax(request, model_class):
    if request.method == 'POST':
        nome = request.POST.get('nome')
        if nome:
            obj, created = model_class.objects.get_or_create(nome=nome)
            return JsonResponse({'id': obj.id, 'nome': obj.nome})
    return JsonResponse({'error': 'Requisição inválida'}, status=400)

@login_required
def add_status_ajax(request): return add_generic_ajax(request, Status)
@login_required
def add_produto_ajax(request): return add_generic_ajax(request, Produto)

class KanbanView(LoginRequiredMixin, TemplateView):
    template_name = 'casos/kanban_board.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = "Kanban de Casos por Etapa"
        
        # Lógica para o novo workflow
        fluxo_principal = FluxoTrabalho.objects.first()
        if not fluxo_principal:
            context['kanban_columns'] = {}
            return context

        casos_filtrados = Caso.objects.filter(etapa_atual__fluxo_trabalho=fluxo_principal)
        # Adicione seus filtros de profissional, status, etc. aqui se necessário

        kanban_columns = {}
        etapas_do_fluxo = fluxo_principal.etapas.order_by('ordem')

        for etapa in etapas_do_fluxo:
            kanban_columns[etapa] = list(casos_filtrados.filter(etapa_atual=etapa))

        context['kanban_columns'] = kanban_columns
        context['profissionais'] = Usuario.objects.filter(is_staff=True)
        context['status_list'] = Status.objects.all()
        return context

@login_required
def update_caso_fase_ajax(request):
    # Esta view precisa ser reescrita para usar 'EtapaFluxo' em vez de 'FaseWorkflow'
    if request.method == 'POST':
        try:
            caso_id = request.POST.get('caso_id')
            nova_etapa_id = request.POST.get('nova_etapa_id')
            caso = get_object_or_404(Caso, pk=caso_id)
            nova_etapa = get_object_or_404(EtapaFluxo, pk=nova_etapa_id)
            
            # Usando a nova função auxiliar para consistência
            _mudar_etapa_fluxo(request, caso, nova_etapa)
            
            messages.success(request, f'Caso #{caso_id} movido para {nova_etapa.nome}.')
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'}, status=400)

class TimesheetListView(LoginRequiredMixin, ListView):
    model = Timesheet
    template_name = 'casos/timesheet_list.html'
    context_object_name = 'timesheets'
    paginate_by = 20
    def get_queryset(self):
        queryset = Timesheet.objects.select_related('caso', 'profissional').order_by('-data_execucao')
        # Adicione seus filtros de profissional, data, etc. aqui
        return queryset
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = "Visão Geral de Timesheet"
        context['profissionais'] = get_user_model().objects.filter(is_staff=True)
        return context

@login_required
def exportar_andamentos_excel(request, caso_pk):
    caso = get_object_or_404(Caso, pk=caso_pk)
    andamentos = caso.andamentos_caso.all()
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="andamentos_caso_{caso.id}.xlsx"'
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Andamentos'
    sheet.merge_cells('A1:C1')
    titulo_cell = sheet['A1']
    titulo_cell.value = f"Relatório de Andamentos - Caso: {caso.titulo_caso}"
    titulo_cell.font = Font(bold=True, size=14)
    titulo_cell.alignment = Alignment(horizontal='center')
    sheet.append([])
    headers = ['Data', 'Usuário', 'Descrição']
    sheet.append(headers)
    for cell in sheet[3]:
        cell.font = Font(bold=True)
    for andamento in andamentos:
        sheet.append([
            andamento.data_andamento.strftime("%d/%m/%Y %H:%M"),
            str(andamento.usuario_criacao) if andamento.usuario_criacao else "Sistema",
            andamento.descricao
        ])
    workbook.save(response)
    return response

@login_required
def exportar_andamentos_pdf(request, caso_pk):
    caso = get_object_or_404(Caso, pk=caso_pk)
    andamentos = caso.andamentos_caso.all()
    logo_url = None
    try:
        logo_config = LogoConfig.objects.get(ativo=True)
        if logo_config.logo:
            logo_url = request.build_absolute_uri(logo_config.logo.url)
    except LogoConfig.DoesNotExist: 
        pass
    context = {'caso': caso, 'andamentos': andamentos, 'logo_url': logo_url}
    html_string = render_to_string('casos/pdf/andamento_pdf_template.html', context)
    html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
    pdf = html.write_pdf()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="andamentos_caso_{caso.id}.pdf"'
    return response

@login_required
def exportar_timesheet_excel(request, caso_pk):
    caso = get_object_or_404(Caso, pk=caso_pk)
    timesheets = caso.timesheets.all().order_by('data_execucao')
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="timesheet_caso_{caso.id}.xlsx"'
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Lançamentos de Horas'
    sheet.merge_cells('A1:D1')
    titulo_cell = sheet['A1']
    titulo_cell.value = f"Relatório de Horas - Caso: {caso.titulo_caso}"
    titulo_cell.font = Font(bold=True, size=14)
    titulo_cell.alignment = Alignment(horizontal='center')
    sheet.row_dimensions[1].height = 20
    headers = ['Data de Execução', 'Profissional', 'Tempo Gasto (HH:MM)', 'Descrição da Atividade']
    sheet.append([])
    sheet.append(headers)
    for cell in sheet[3]:
        cell.font = Font(bold=True)
    sheet.column_dimensions['A'].width, sheet.column_dimensions['B'].width, sheet.column_dimensions['C'].width, sheet.column_dimensions['D'].width = 18, 30, 20, 60
    for ts in timesheets:
        data_formatada = ts.data_execucao.strftime("%d/%m/%Y")
        horas, minutos = ts.tempo.seconds // 3600, (ts.tempo.seconds % 3600) // 60
        tempo_formatado = f"{horas:02d}:{minutos:02d}"
        sheet.append([
            data_formatada,
            ts.profissional.get_full_name() or ts.profissional.username,
            tempo_formatado,
            ts.descricao
        ])
    if timesheets:
        proxima_linha = sheet.max_row + 1
        total_label_cell = sheet.cell(row=proxima_linha, column=2)
        total_label_cell.value, total_label_cell.font, total_label_cell.alignment = "Total de Horas:", Font(bold=True), Alignment(horizontal='right')
        total_value_cell = sheet.cell(row=proxima_linha, column=3)
        total_value_cell.value, total_value_cell.font = caso.total_horas_trabalhadas, Font(bold=True)
    workbook.save(response)
    return response

    
@login_required
def enviar_timesheet_email(request, caso_pk):
    caso = get_object_or_404(Caso, pk=caso_pk)
    contexto = {'caso': caso, 'usuario_acao': request.user}
    sucesso, mensagem = enviar_notificacao(slug_evento='envio_relatorio_timesheet', contexto=contexto)
    if sucesso:
        messages.success(request, mensagem)
    else:
        messages.error(request, mensagem)
    return redirect(reverse('casos:caso_detail', kwargs={'pk': caso_pk}) + '#timesheet-tab-pane')

@login_required
def add_timesheet(request, caso_pk):
    caso = get_object_or_404(models.Caso, pk=caso_pk)
    if request.method == 'POST':
        form = forms.TimesheetForm(request.POST)
        if form.is_valid():
            timesheet = form.save(commit=False)
            timesheet.caso = caso
            # O campo 'tempo' já vem formatado corretamente do formulário
            timesheet.save()
            messages.success(request, "Lançamento de timesheet adicionado com sucesso.")
        else:
            # Se o formulário for inválido, mostre os erros
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Erro no campo '{field}': {error}")

    return redirect(reverse('casos:caso_detail', kwargs={'pk': caso_pk}) + '#timesheet-tab-pane')

@login_required
def delete_timesheet(request, ts_pk):
    timesheet = get_object_or_404(Timesheet, pk=ts_pk)
    caso_pk = timesheet.caso.pk
    if request.method == 'POST':
        timesheet.delete()
    return redirect(reverse('casos:caso_detail', kwargs={'pk': caso_pk}) + '#timesheet-tab-pane')

def add_generic_ajax(request, model_class):
    if request.method == 'POST':
        nome = request.POST.get('nome')
        if nome:
            obj, created = model_class.objects.get_or_create(nome=nome)
            return JsonResponse({'id': obj.id, 'nome': obj.nome})
    return JsonResponse({'error': 'Requisição inválida'}, status=400)

@login_required
def add_status_ajax(request):
    from .models import Status
    return add_generic_ajax(request, Status)

@login_required
def add_analista_ajax(request):
    from .models import Analista
    return add_generic_ajax(request, Analista)

@login_required
def add_produto_ajax(request):
    from .models import Produto
    return add_generic_ajax(request, Produto)

@login_required
def executar_acao(request, acao_pk, opcao_pk=None):
    instancia_acao = get_object_or_404(models.InstanciaAcao, pk=acao_pk)
    caso = instancia_acao.caso
    opcao_decisao = None
    label_decisao = "Concluída"

    if opcao_pk:
        opcao_decisao = get_object_or_404(models.OpcaoDecisao, pk=opcao_pk)
        label_decisao = opcao_decisao.label_do_botao

    if request.method == 'POST':
        descricao_conclusao = request.POST.get('descricao_conclusao', '')
        
        instancia_acao.status = 'C'
        instancia_acao.data_conclusao = timezone.now()
        instancia_acao.descricao_conclusao = descricao_conclusao
        instancia_acao.save()
        
        models.FluxoInterno.objects.create(caso=caso, data_fluxo=timezone.now().date(), descricao=f"[AÇÃO] '{instancia_acao.acao_modelo.titulo}' concluída com decisão '{label_decisao}'.\n{descricao_conclusao}", usuario_criacao=request.user)

        if opcao_decisao:
            if opcao_decisao.avancar_proxima_etapa:
                etapa_atual = caso.etapa_atual
                if etapa_atual:
                    proxima_etapa = models.EtapaFluxo.objects.filter(fluxo_trabalho=etapa_atual.fluxo_trabalho, ordem__gt=etapa_atual.ordem).order_by('ordem').first()
                    _mudar_etapa_fluxo(request, caso, proxima_etapa)
                    if not proxima_etapa: messages.success(request, "Fluxo de trabalho finalizado.")
            elif opcao_decisao.mudar_etapa_para:
                _mudar_etapa_fluxo(request, caso, opcao_decisao.mudar_etapa_para)
            
            if opcao_decisao.criar_nova_acao:
                models.InstanciaAcao.objects.create(caso=caso, acao_modelo=opcao_decisao.criar_nova_acao, responsavel=request.user)
            
            if opcao_decisao.atualizar_status_caso:
                caso.status = opcao_decisao.atualizar_status_caso
                caso.save(update_fields=['status'])
            
            # ... (Lógica de e-mail e aguardar dias)

        return redirect(reverse('casos:caso_detail', kwargs={'pk': caso.pk}) + '#acoes-tab-pane')

    return redirect('casos:caso_detail', kwargs={'pk': caso.pk})


@login_required
def reabrir_acao(request, acao_instancia_pk):
    acao_instancia = get_object_or_404(models.InstanciaAcao, pk=acao_instancia_pk)
    if request.method == 'POST':
        acao_instancia.status = 'P'
        acao_instancia.data_conclusao = None
        acao_instancia.save()
        messages.success(request, f"Ação '{acao_instancia.acao_modelo.titulo}' reaberta.")
    return redirect(reverse('casos:caso_detail', kwargs={'pk': acao_instancia.caso.pk}) + '#acoes-tab-pane')

@login_required
def deletar_acao(request, acao_instancia_pk):
    acao_instancia = get_object_or_404(models.InstanciaAcao, pk=acao_instancia_pk)
    caso_pk = acao_instancia.caso.pk
    if request.method == 'POST':
        acao_instancia.delete()
        messages.success(request, f"Ação '{acao_instancia.acao_modelo.titulo}' deletada.")
    return redirect(reverse('casos:caso_detail', kwargs={'pk': caso_pk}) + '#acoes-tab-pane')
@login_required
def deletar_acao(request, acao_instancia_pk):
    acao_instancia = get_object_or_404(InstanciaAcao, pk=acao_instancia_pk)
    caso_pk = acao_instancia.caso.pk
    if request.method == 'POST':
        acao_instancia.delete()
        messages.success(request, f"Ação '{acao_instancia.acao_modelo.titulo}' deletada.")
    return redirect(reverse('casos:caso_detail', kwargs={'pk': caso_pk}) + '#acoes-tab-pane')

@login_required
def add_timesheet(request, caso_pk):
    caso = get_object_or_404(models.Caso, pk=caso_pk)
    if request.method == 'POST':
        # Agora usamos o novo formulário inteligente
        form = LancamentoHorasForm(request.POST)
        if form.is_valid():
            lancamento = form.save(commit=False)
            lancamento.caso = caso
            lancamento.save() # O método save() do formulário já faz a conversão do tempo
            messages.success(request, "Lançamento de horas adicionado com sucesso.")
        else:
            messages.error(request, f"Erro no formulário: {form.errors.as_text()}")
    return redirect(reverse('casos:caso_detail', kwargs={'pk': caso_pk}) + '#timesheet-tab-pane')



@login_required
def add_lancamento_horas(request, caso_pk):
    caso = get_object_or_404(models.Caso, pk=caso_pk)
    if request.method == 'POST':
        form = LancamentoHorasForm(request.POST)
        if form.is_valid():
            lancamento = form.save(commit=False)
            lancamento.caso = caso
            
            tempo_str = form.cleaned_data['tempo_str']
            horas, minutos = map(int, tempo_str.split(':'))
            lancamento.minutos_gastos = (horas * 60) + minutos
            
            lancamento.save()
            messages.success(request, "Lançamento adicionado.")
        else:
            messages.error(request, f"Erro: {form.errors.as_text()}")
    return redirect(reverse('casos:caso_detail', kwargs={'pk': caso_pk}) + '#timesheet-tab-pane')

class LancamentoHorasUpdateView(LoginRequiredMixin, UpdateView):
    model = models.Timesheet
    form_class = forms.LancamentoHorasForm
    template_name = 'casos/lancamento_horas_form.html'

    def get_form_kwargs(self):
        """
        Este método é o lugar certo para adicionar valores iniciais
        sem interferir com os dados da instância.
        """
        kwargs = super().get_form_kwargs()
        
        # Se estamos editando um objeto existente...
        if self.object and self.object.minutos_gastos is not None:
            # Criamos um dicionário 'initial' se ele não existir
            if 'initial' not in kwargs:
                kwargs['initial'] = {}
            
            # Calculamos e adicionamos o valor do tempo_str
            horas = self.object.minutos_gastos // 60
            minutos = self.object.minutos_gastos % 60
            kwargs['initial']['tempo_str'] = f"{horas:02d}:{minutos:02d}"
            
        return kwargs

    def form_valid(self, form):
        lancamento = form.save(commit=False)
        tempo_str = form.cleaned_data['tempo_str']
        horas, minutos = map(int, tempo_str.split(':'))
        lancamento.minutos_gastos = (horas * 60) + minutos
        lancamento.save()
        messages.success(self.request, "Lançamento atualizado com sucesso.")
        return redirect(self.get_success_url())

    def get_success_url(self):
        return reverse_lazy('casos:caso_detail', kwargs={'pk': self.object.caso.pk}) + '#timesheet-tab-pane'
    

# ==============================================================
# NOVAS VIEWS PARA GERENCIAR DESPESAS
# ==============================================================

class DespesaCreateView(LoginRequiredMixin, CreateView):
    model = DespesaCaso
    form_class = DespesaCasoForm
    template_name = 'casos/despesa_form.html'

    def form_valid(self, form):
        # Pega o caso da URL e o associa à nova despesa antes de salvar
        caso = get_object_or_404(Caso, pk=self.kwargs['caso_pk'])
        form.instance.caso = caso
        messages.success(self.request, "Despesa adicionada com sucesso!")
        return super().form_valid(form)

    def get_success_url(self):
        # Durante a criação, o 'caso_pk' está na URL, não no objeto.
        # Nós pegamos o pk da URL para garantir que o redirecionamento funcione.
        
        # --- A CORREÇÃO ESTÁ AQUI ---
        # 1. Primeiro, pegamos a variável 'caso_pk' da URL
        caso_pk_da_url = self.kwargs.get('caso_pk')
        
        # 2. Agora, usamos essa variável para construir a URL de retorno
        return reverse_lazy('casos:caso_detail', kwargs={'pk': caso_pk_da_url}) + '#despesas-tab-pane'
    
class DespesaUpdateView(LoginRequiredMixin, UpdateView):
    model = DespesaCaso
    form_class = DespesaCasoForm
    template_name = 'casos/despesa_form.html'

    def form_valid(self, form):
        messages.success(self.request, "Despesa atualizada com sucesso!")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('casos:caso_detail', kwargs={'pk': self.object.caso.pk}) + '#despesas-tab-pane'

class DespesaDeleteView(LoginRequiredMixin, DeleteView):
    model = DespesaCaso
    template_name = 'casos/despesa_confirm_delete.html'

    def form_valid(self, form):
        messages.success(self.request, "Despesa excluída com sucesso!")
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse_lazy('casos:caso_detail', kwargs={'pk': self.object.caso.pk}) + '#despesas-tab-pane'
    
@login_required
def exportar_despesas_excel(request, caso_pk):
    caso = get_object_or_404(Caso, pk=caso_pk)
    despesas = caso.despesas.all().order_by('data_despesa')
    total_despesas = despesas.aggregate(total=Sum('valor'))['total'] or 0.00

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="despesas_caso_{caso.id}.xlsx"'
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Despesas do Caso'

    # Título do Relatório
    sheet.merge_cells('A1:D1')
    titulo_cell = sheet['A1']
    titulo_cell.value = f"Relatório de Despesas - Caso: {caso.titulo_caso}"
    titulo_cell.font = Font(bold=True, size=14)
    titulo_cell.alignment = Alignment(horizontal='center')
    sheet.row_dimensions[1].height = 20
    sheet.append([]) # Linha em branco

    # Cabeçalhos da Tabela
    headers = ['Data da Despesa', 'Descrição', 'Valor (R$)', '']
    sheet.append(headers)
    for cell in sheet[3]:
        cell.font = Font(bold=True)
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 60
    sheet.column_dimensions['C'].width = 15

    # Dados
    for despesa in despesas:
        sheet.append([
            despesa.data_despesa,
            despesa.descricao,
            despesa.valor,
        ])
        sheet.cell(row=sheet.max_row, column=1).number_format = 'DD/MM/YYYY'
        sheet.cell(row=sheet.max_row, column=3).number_format = '"R$" #,##0.00'

    # Linha do Total
    proxima_linha = sheet.max_row + 2
    sheet.cell(row=proxima_linha, column=2, value="Total Geral:").font = Font(bold=True)
    sheet.cell(row=proxima_linha, column=2, value="Total Geral:").alignment = Alignment(horizontal='right')
    total_cell = sheet.cell(row=proxima_linha, column=3, value=total_despesas)
    total_cell.font = Font(bold=True)
    total_cell.number_format = '"R$" #,##0.00'

    workbook.save(response)
    return response


@login_required
def exportar_despesas_pdf(request, caso_pk):
    # --- Importações locais ---
    from django.http import HttpResponse
    from django.db.models import Sum
    from django.utils import timezone
    from weasyprint import HTML
    from .models import Caso
    # ---------------------------

    try:
        # Busca os dados do banco
        caso = get_object_or_404(Caso, pk=caso_pk)
        despesas = caso.despesas.all().order_by('data_despesa')
        total_despesas = despesas.aggregate(total=Sum('valor'))['total'] or 0.00
        data_emissao = timezone.now()

        # --- Construção manual do HTML ---
        despesas_html_rows = ""
        if despesas.exists():
            for despesa in despesas:
                # Formatando os valores para exibição
                data_formatada = despesa.data_despesa.strftime('%d/%m/%Y')
                valor_formatado = f"R$ {despesa.valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                
                despesas_html_rows += f"""
                    <tr>
                        <td>{data_formatada}</td>
                        <td>{despesa.descricao}</td>
                        <td class="text-end">{valor_formatado}</td>
                    </tr>
                """
        else:
            despesas_html_rows = '<tr><td colspan="3" style="text-align: center;">Nenhuma despesa registrada.</td></tr>'

        # Formata o total
        total_formatado = f"R$ {total_despesas:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

        # Template HTML principal
        html_string = f"""
        <!DOCTYPE html>
        <html lang="pt-br">
        <head>
            <meta charset="UTF-8">
            <title>Relatório de Despesas - Caso #{caso.id}</title>
            <style>
                @page {{ size: A4; margin: 1.5cm; }}
                body {{ font-family: 'Helvetica', sans-serif; font-size: 10pt; color: #333; }}
                .header {{ text-align: center; margin-bottom: 20px; border-bottom: 1px solid #ccc; padding-bottom: 10px; }}
                .header h1 {{ margin: 0; font-size: 18pt; }}
                .header h2 {{ margin: 5px 0 0 0; font-size: 12pt; color: #666; }}
                .info-caso {{ margin-bottom: 20px; }}
                .info-caso p {{ margin: 2px 0; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 15px; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; font-size: 11pt; }}
                .text-end {{ text-align: right; }}
                .footer {{ position: fixed; bottom: -1cm; left: 0; right: 0; text-align: center; font-size: 8pt; color: #888; }}
                .total-row {{ font-weight: bold; background-color: #f9f9f9; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>Relatório de Despesas</h1>
                <h2>Caso: {caso.titulo_caso or f'#{caso.id}'}</h2>
            </div>
            <div class="info-caso">
                <p><strong>Cliente:</strong> {caso.cliente.nome_razao_social}</p>
                <p><strong>Produto:</strong> {caso.produto.nome}</p>
                <p><strong>Data de Emissão:</strong> {data_emissao.strftime('%d/%m/%Y %H:%M')}</p>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>Data</th>
                        <th>Descrição</th>
                        <th class="text-end">Valor</th>
                    </tr>
                </thead>
                <tbody>
                    {despesas_html_rows}
                </tbody>
                <tfoot>
                    <tr class="total-row">
                        <td colspan="2" class="text-end">TOTAL GERAL:</td>
                        <td class="text-end">{total_formatado}</td>
                    </tr>
                </tfoot>
            </table>
            <div class="footer">
                Gerado por Sistema RCA em {data_emissao.strftime('%d/%m/%Y')}
            </div>
        </body>
        </html>
        """
        # ---------------------------------
        
        html = HTML(string=html_string)
        pdf = html.write_pdf()
        
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="despesas_caso_{caso.id}.pdf"'
        
        return response

    except Exception as e:
        # Em caso de qualquer erro inesperado, retorna uma resposta de erro simples
        print(f"ERRO FATAL ao gerar PDF de despesas: {e}")
        return HttpResponse(f"Ocorreu um erro ao gerar o PDF: {e}", status=500)
    

class AcordoCreateView(LoginRequiredMixin, CreateView):
    model = AcordoCaso
    form_class = AcordoCasoForm
    template_name = 'casos/acordo_form.html'

    def form_valid(self, form):
        caso = get_object_or_404(Caso, pk=self.kwargs['caso_pk'])
        form.instance.caso = caso
        messages.success(self.request, "Acordo adicionado com sucesso! As parcelas foram geradas automaticamente.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('casos:caso_detail', kwargs={'pk': self.kwargs['caso_pk']}) + '#acordos-tab-pane'

class AcordoUpdateView(LoginRequiredMixin, UpdateView):
    model = AcordoCaso
    form_class = AcordoCasoForm
    template_name = 'casos/acordo_form.html'

    def form_valid(self, form):
        messages.success(self.request, "Acordo atualizado com sucesso!")
        # ATENÇÃO: A edição de um acordo NÃO recria as parcelas. Isso evita perda de dados.
        # Se precisar recriar, a lógica seria mais complexa (ex: apagar parcelas antigas e criar novas).
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('casos:caso_detail', kwargs={'pk': self.object.caso.pk}) + '#acordos-tab-pane'

class AcordoDeleteView(LoginRequiredMixin, DeleteView):
    model = AcordoCaso
    template_name = 'casos/acordo_confirm_delete.html'
    
    def form_valid(self, form):
        messages.success(self.request, "Acordo e todas as suas parcelas foram excluídos com sucesso!")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('casos:caso_detail', kwargs={'pk': self.object.caso.pk}) + '#acordos-tab-pane'


@login_required
def quitar_parcela(request, pk):
    # Encontra a parcela específica ou retorna um erro 404
    parcela = get_object_or_404(ParcelaAcordo, pk=pk)
    caso = parcela.acordo.caso

    # Ação só deve ocorrer se o método for POST (enviado pelo formulário do botão)
    if request.method == 'POST':
        # 1. Atualiza a parcela
        parcela.situacao = 'quitado'
        parcela.data_quitacao = timezone.now().date()
        parcela.save()

        # 2. Cria o registro no Fluxo Interno
        FluxoInternoModel.objects.create(
            caso=caso,
            data_fluxo=timezone.now().date(),
            descricao=(
                f"[ACORDO] Parcela #{parcela.numero_parcela} "
                f"(Venc: {parcela.data_vencimento.strftime('%d/%m/%Y')}) "
                f"do acordo de {parcela.acordo.data_acordo.strftime('%d/%m/%Y')} foi quitada."
            ),
            usuario_criacao=request.user
        )
        
        messages.success(request, f"Parcela #{parcela.numero_parcela} quitada com sucesso!")

    # 3. Redireciona o usuário de volta para a página do caso, na aba de acordos
    return redirect(reverse('casos:caso_detail', kwargs={'pk': caso.pk}) + '#acordos-tab-pane')


@login_required
def exportar_acordos_excel(request, caso_pk):
    caso = get_object_or_404(Caso, pk=caso_pk)
    # Usamos prefetch_related para otimizar a busca das parcelas
    acordos = caso.acordos.prefetch_related('parcelas').order_by('data_acordo')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="acordos_caso_{caso.id}.xlsx"'
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Acordos e Parcelas'

    # Estilos para a planilha
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    acordo_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    center_align = Alignment(horizontal='center')

    # Título do Relatório (agora com 7 colunas)
    sheet.merge_cells('A1:G1')
    titulo_cell = sheet['A1']
    titulo_cell.value = f"Relatório de Acordos - Caso: {caso.titulo_caso or f'#{caso.id}'}"
    titulo_cell.font = Font(bold=True, size=16)
    titulo_cell.alignment = center_align
    sheet.row_dimensions[1].height = 25
    sheet.append([]) # Linha em branco

    # Cabeçalhos da Tabela (com a nova coluna de Valor)
    headers = [
        'Acordo', 
        'Descrição', 
        'Nº da Parcela', 
        'Valor da Parcela (R$)', # <<< COLUNA ADICIONADA
        'Vencimento', 
        'Situação', 
        'Data de Quitação'
    ]
    sheet.append(headers)
    for cell in sheet[3]:
        cell.font = header_font
        cell.fill = header_fill

    # Definindo a largura das colunas
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 40
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 20 # <<< LARGURA PARA A NOVA COLUNA
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['G'].width = 18

    # Preenchendo os dados
    if not acordos:
        sheet.append(["Nenhum acordo encontrado para este caso."])
    else:
        for acordo in acordos:
            # Linha de cabeçalho para cada acordo
            acordo_header = [f"Acordo de {acordo.data_acordo.strftime('%d/%m/%Y')}", acordo.descricao, "", "", "", "", ""]
            sheet.append(acordo_header)
            acordo_row = sheet.max_row
            sheet.merge_cells(start_row=acordo_row, start_column=1, end_row=acordo_row, end_column=7) # <<< AJUSTADO PARA 7 COLUNAS
            for cell in sheet[acordo_row]:
                cell.font = Font(bold=True)
                cell.fill = acordo_fill

            # Parcelas do acordo (com o novo campo de valor)
            for parcela in acordo.parcelas.all():
                sheet.append([
                    "", 
                    "",
                    parcela.numero_parcela,
                    acordo.valor_parcela, # <<< VALOR DA PARCELA ADICIONADO AQUI
                    parcela.data_vencimento,
                    parcela.get_situacao_display(),
                    parcela.data_quitacao if parcela.data_quitacao else "",
                ])
                # Formatando as células da linha que acabamos de adicionar
                row_idx = sheet.max_row
                sheet.cell(row=row_idx, column=4).number_format = '"R$" #,##0.00' # Formato de moeda
                sheet.cell(row=row_idx, column=5).number_format = 'DD/MM/YYYY'   # Formato de data
                sheet.cell(row=row_idx, column=7).number_format = 'DD/MM/YYYY'   # Formato de data

    workbook.save(response)
    return response


@login_required
def exportar_acordos_pdf(request, caso_pk):
    # --- Importações locais ---
    from django.http import HttpResponse
    from django.utils import timezone
    from weasyprint import HTML
    from .models import Caso
    # ---------------------------

    try:
        # Busca os dados do banco
        caso = get_object_or_404(Caso, pk=caso_pk)
        acordos = caso.acordos.prefetch_related('parcelas').order_by('data_acordo')
        data_emissao = timezone.now()

        # --- Construção do HTML com Estilo ---
        
        # Bloco de estilo CSS (o mesmo do seu template original)
        css_style = """
            @page { size: A4; margin: 1.5cm; }
            body { font-family: 'Helvetica', sans-serif; font-size: 10pt; color: #333; }
            .header { text-align: center; margin-bottom: 20px; border-bottom: 1px solid #ccc; padding-bottom: 10px; }
            .header img { max-height: 70px; max-width: 250px; margin-bottom: 10px; }
            .header h1 { margin: 0; font-size: 18pt; }
            .header h2 { margin: 5px 0 0 0; font-size: 12pt; color: #666; font-weight: normal; }
            .info-caso { margin-bottom: 20px; padding: 10px; background-color: #f9f9f9; border: 1px solid #eee; border-radius: 5px; }
            .info-caso p { margin: 4px 0; }
            .acordo-bloco { margin-bottom: 25px; page-break-inside: avoid; }
            .acordo-header { background-color: #f2f2f2; padding: 8px; border: 1px solid #ddd; border-bottom: none; }
            .acordo-header h3 { margin: 0; font-size: 12pt; }
            .acordo-header small { color: #555; }
            .acordo-descricao { font-style: italic; padding: 10px; border: 1px solid #ddd; border-top: none; border-bottom: none; }
            table { width: 100%; border-collapse: collapse; }
            th, td { border: 1px solid #ddd; padding: 6px; text-align: left; font-size: 9pt; }
            th { background-color: #fafafa; }
            .text-end { text-align: right; }
            .text-center { text-align: center; }
            .badge { display: inline-block; padding: .25em .6em; font-size: 75%; font-weight: 700; line-height: 1; text-align: center; white-space: nowrap; vertical-align: baseline; border-radius: .25rem; color: #fff; }
            .bg-success { background-color: #198754; }
            .bg-warning { background-color: #ffc107; color: #000 !important; }
            .currency { white-space: nowrap; }
            .footer { position: fixed; bottom: -1cm; left: 0; right: 0; text-align: center; font-size: 8pt; color: #888; }
        """

        # Corpo do HTML
        body_html = f"""
        <div class="header">
            <h1>Relatório de Acordos</h1>
            <h2>Caso: {caso.titulo_caso or f'#{caso.id}'}</h2>
        </div>
        <div class="info-caso">
            <p><strong>Cliente:</strong> {caso.cliente.nome_razao_social}</p>
            <p><strong>Produto:</strong> {caso.produto.nome}</p>
            <p><strong>Data de Emissão:</strong> {data_emissao.strftime('%d/%m/%Y %H:%M')}</p>
        </div>
        """

        # Loop para gerar cada bloco de acordo
        if acordos.exists():
            for acordo in acordos:
                # Formatações
                data_acordo_fmt = acordo.data_acordo.strftime('%d/%m/%Y')
                valor_parcela_fmt = f"R$ {acordo.valor_parcela:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                valor_total_fmt = f"R$ {acordo.valor_total:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                
                # Cabeçalho do acordo
                body_html += f"""
                <div class="acordo-bloco">
                    <div class="acordo-header">
                        <h3>Acordo de {data_acordo_fmt}</h3>
                        <small>{acordo.quantidade_parcelas} parcela(s) de {valor_parcela_fmt} | Total: {valor_total_fmt}</small>
                    </div>
                """
                if acordo.descricao:
                    body_html += f'<div class="acordo-descricao"><p>"{acordo.descricao}"</p></div>'
                
                # Tabela de parcelas
                body_html += """
                <table>
                    <thead>
                        <tr>
                            <th style="width: 5%;">#</th>
                            <th style="width: 20%;">Vencimento</th>
                            <th class="text-end" style="width: 20%;">Valor</th>
                            <th style="width: 20%;">Situação</th>
                            <th style="width: 25%;">Data de Quitação</th>
                        </tr>
                    </thead>
                    <tbody>
                """
                for parcela in acordo.parcelas.all():
                    vencimento_fmt = parcela.data_vencimento.strftime('%d/%m/%Y')
                    quitacao_fmt = parcela.data_quitacao.strftime('%d/%m/%Y') if parcela.data_quitacao else '-'
                    badge_class = "bg-success" if parcela.situacao == 'quitado' else "bg-warning"
                    
                    body_html += f"""
                    <tr>
                        <td class="text-center">{parcela.numero_parcela}</td>
                        <td>{vencimento_fmt}</td>
                        <td class="text-end currency">{valor_parcela_fmt}</td>
                        <td><span class="badge {badge_class}">{parcela.get_situacao_display()}</span></td>
                        <td>{quitacao_fmt}</td>
                    </tr>
                    """
                body_html += "</tbody></table></div>"
        else:
            body_html += "<p>Nenhum acordo registrado para este caso.</p>"

        # Template HTML final
        html_string = f"""
        <!DOCTYPE html>
        <html lang="pt-br">
        <head>
            <meta charset="UTF-8">
            <title>Relatório de Acordos - Caso #{caso.id}</title>
            <style>{css_style}</style>
        </head>
        <body>
            {body_html}
            <div class="footer">Gerado por Sistema RCA em {data_emissao.strftime('%d/%m/%Y')}</div>
        </body>
        </html>
        """
        
        html = HTML(string=html_string)
        pdf = html.write_pdf()
        
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="acordos_caso_{caso.id}.pdf"'
        
        return response

    except Exception as e:
        return HttpResponse(f"Ocorreu um erro ao gerar o PDF: {e}", status=500)